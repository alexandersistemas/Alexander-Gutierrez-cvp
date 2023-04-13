from apiproduct.models import Producto
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def reporte_productos_to_xlsx(request):
   Productos = Producto.objects.all()

   workbook = Workbook()
   worksheet = workbook.active
   worksheet.title = 'Productos'

   columns = [
      'Producto',
      'Descripcion',
      'Precio',
      'Cantidad'
   ]
   row_num = 1

   for col_num, column_title in enumerate(columns, 1):
      cell = worksheet.cell(row=row_num, column=col_num)
      cell.value = column_title

   contador = 2
   for product in Productos:
      column = str(contador)
      worksheet['A' + column] = product.nombre
      worksheet['B' + column] = product.descripcion
      worksheet['C' + column] = product.precio
      worksheet['D' + column] = product.cantidad
      contador += 1

   nombre_archivo = "ListadoProductos.xlsx"
   response = HttpResponse(content_type="application/ms-excel")
   contenido = "attachment; filename = {0}".format(nombre_archivo)
   response["Content-Disposition"] = contenido
   workbook.save(response)
   return response


def login(request):
   return render(request, "login.html")

def validarLogin(request):
   if request.method == "POST":

      username = request.POST.get('username')
      password = request.POST.get('password')
      print('USERNAMEEEE ' +str(username))
      print('passwordddd ' +str(password))

      user = authenticate(username=username.strip(), password=password.strip())

      if user is not None:
         productos = Producto.objects.all()
         return render(request, "index.html", context={'productos': productos})
      else:
         error_message = 'Credenciales incorrectas!'
         return render(request, "login.html", context={'msg': error_message})



@login_required
def home(request):
   productos = Producto.objects.all()
   return render(request, "index.html", context={'productos':productos})

def detail(request, id):
   producto = Producto.objects.get(id=id)
   return render(request, "detail.html", context={'producto':producto})

def adicionar(request):
    return render(request, "Adicionar.html")

def borrado(request, id):
   producto = Producto.objects.get(id=id)
   return render(request, "Borrado.html", context={'producto':producto})

def ActualizaProducto(request):
   if request.method == "POST":
      nuevaCantidad = request.POST.get('cantidad')
      postId = request.POST.get('postId')
      Producto.objects.filter(id=postId).update(cantidad=nuevaCantidad)

   productos = Producto.objects.all()
   return render(request, "index.html", context={'productos': productos})

def GuardaProducto(request):
   if request.method == "POST":
      nuevoNombre = request.POST.get('nombre')
      nuevoPrecio = request.POST.get('precio')
      nuevaDescripcion = request.POST.get('descripcion')
      nuevaCantidad = request.POST.get('cantidad')
      a = Producto(nombre = nuevoNombre, precio = nuevoPrecio, descripcion = nuevaDescripcion, cantidad = nuevaCantidad)
      a.save()

   productos = Producto.objects.all()
   return render(request, "index.html", context={'productos': productos})

def BorrarProducto(request):
   if request.method == "POST":
      postId = request.POST.get('postId')
      Producto.objects.filter(id=postId).delete()

   productos = Producto.objects.all()
   return render(request, "index.html", context={'productos': productos})
